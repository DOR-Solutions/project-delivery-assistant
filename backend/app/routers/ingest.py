import csv
import io
from pathlib import Path
from typing import Any
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from pydantic import BaseModel
from sqlalchemy.orm import Session
from ..database import get_db
from .. import models, ingest_watch
from .. import portfolio
from openpyxl import load_workbook
from datetime import datetime

router = APIRouter(prefix="/api/ingest", tags=["ingest"])


class DailyIngestPayload(BaseModel):
    project_id: str
    source_ref: str = ""
    bag_days: list[dict[str, Any]] | None = None
    risks: list[dict[str, Any]] | None = None
    report_snapshots: list[dict[str, Any]] | None = None
    work_log: list[dict[str, Any]] | None = None
    milestones: list[dict[str, Any]] | None = None


class DailyIngestResult(BaseModel):
    project_id: str
    source_ref: str
    stored: dict[str, int]


def _normalize_upload_rows(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    datasets: dict[str, list[dict[str, Any]]] = {"bag_days": [], "risks": [], "report_snapshots": [], "work_log": [], "milestones": []}
    for row in rows:
        if not row:
            continue
        normalized = {str(k).strip(): (v if v is not None else "") for k, v in row.items() if k is not None}
        kind = str(normalized.get("kind", "")).strip().lower()
        if kind in {"bag_days", "bagday", "bag-day", "bag day"}:
            datasets["bag_days"].append(normalized)
            continue
        if kind in {"risks", "risk"}:
            datasets["risks"].append(normalized)
            continue
        if kind in {"report_snapshots", "report", "snapshot", "report_snapshot"}:
            datasets["report_snapshots"].append(normalized)
            continue
        if kind in {"work_log", "worklog", "work-log", "work"}:
            datasets["work_log"].append(normalized)
            continue
        if kind in {"milestones", "milestone"}:
            datasets["milestones"].append(normalized)
            continue
        if "series" in normalized and "planned" in normalized and "actual" in normalized:
            datasets["bag_days"].append(normalized)
        elif "title" in normalized and ("likelihood" in normalized or "impact" in normalized or "status" in normalized):
            datasets["risks"].append(normalized)
        elif "completion_pct" in normalized or "rag" in normalized:
            datasets["report_snapshots"].append(normalized)
        elif "activity" in normalized or "contractor" in normalized or "pct" in normalized:
            datasets["work_log"].append(normalized)
        elif "title" in normalized and ("detail" in normalized or "on_track" in normalized):
            datasets["milestones"].append(normalized)
    return datasets


def _parse_upload_file(upload: UploadFile) -> dict[str, list[dict[str, Any]]]:
    if upload.filename is None:
        raise HTTPException(400, "Upload requires a filename")
    suffix = Path(upload.filename).suffix.lower()
    content = upload.file.read()
    if suffix == ".csv":
        text = content.decode("utf-8-sig")
        rows = list(csv.DictReader(io.StringIO(text)))
    elif suffix == ".xlsx":
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows: list[dict[str, Any]] = []
        headers = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx == 0:
                headers = [str(cell) if cell is not None else "" for cell in row]
                continue
            row_data = {headers[i]: row[i] if i < len(row) else "" for i in range(len(headers))}
            rows.append(row_data)
    else:
        raise HTTPException(400, "Only CSV and XLSX uploads are supported")
    return _normalize_upload_rows(rows)


def _upsert_bag_days(db: Session, project_id: str, source_ref: str, rows: list[dict[str, Any]]) -> int:
    stored = 0
    for row in rows or []:
        date_value = row.get("date")
        series = row.get("series") or "throughput"
        if not date_value:
            continue
        existing = db.query(models.BagDay).filter(
            models.BagDay.project_id == project_id,
            models.BagDay.date == date_value,
            models.BagDay.series == series,
        ).first()
        if existing:
            existing.planned = row.get("planned", existing.planned)
            existing.actual = row.get("actual", existing.actual)
            existing.capacity = row.get("capacity", existing.capacity)
            existing.day_type = row.get("day_type", existing.day_type)
            existing.breakdown = row.get("breakdown", existing.breakdown or {})
            existing.source_ref = source_ref
            existing.created_at = existing.created_at or datetime.utcnow()
        else:
            db.add(models.BagDay(project_id=project_id, date=date_value, day=row.get("day", ""),
                                 series=series, planned=row.get("planned", 0), actual=row.get("actual", 0),
                                 capacity=row.get("capacity", 0), day_type=row.get("day_type", ""),
                                 breakdown=row.get("breakdown") or {}, source_ref=source_ref))
        stored += 1
    return stored


def _upsert_risks(db: Session, project_id: str, source_ref: str, rows: list[dict[str, Any]]) -> int:
    stored = 0
    for row in rows or []:
        title = row.get("title")
        if not title:
            continue
        existing = db.query(models.Risk).filter(models.Risk.project_id == project_id, models.Risk.title == title).first()
        if existing:
            existing.area = row.get("area", existing.area)
            existing.likelihood = row.get("likelihood", existing.likelihood)
            existing.impact = row.get("impact", existing.impact)
            existing.status = row.get("status", existing.status)
            existing.mitigation = row.get("mitigation", existing.mitigation)
            existing.owner = row.get("owner", existing.owner)
            existing.updated_at = datetime.utcnow()
            existing.source_ref = source_ref
        else:
            risk_id = f"{project_id}-{title}".replace(" ", "-").replace("/", "-").lower()
            db.add(models.Risk(id=risk_id, project_id=project_id, title=title, area=row.get("area", ""),
                               likelihood=row.get("likelihood", 0), impact=row.get("impact", 0),
                               mitigation=row.get("mitigation", ""), owner=row.get("owner", ""),
                               status=row.get("status", "open"), source_ref=source_ref))
        stored += 1
    return stored


def _upsert_report_snapshots(db: Session, project_id: str, source_ref: str, rows: list[dict[str, Any]]) -> int:
    stored = 0
    for row in rows or []:
        date_value = row.get("date")
        if not date_value:
            continue
        existing = db.query(models.ReportSnapshot).filter(models.ReportSnapshot.project_id == project_id,
                                                         models.ReportSnapshot.date == date_value).first()
        if existing:
            existing.completion_pct = row.get("completion_pct", existing.completion_pct)
            existing.rag = row.get("rag", existing.rag)
            existing.data = row.get("data", existing.data or {})
            existing.source_ref = source_ref
        else:
            db.add(models.ReportSnapshot(project_id=project_id, date=date_value,
                                         completion_pct=row.get("completion_pct", 0), rag=row.get("rag", "amber"),
                                         data=row.get("data") or {}, source_ref=source_ref))
        stored += 1
    return stored


def _upsert_work_log(db: Session, project_id: str, source_ref: str, rows: list[dict[str, Any]]) -> int:
    stored = 0
    for row in rows or []:
        date_value = row.get("date")
        activity = row.get("activity")
        if not date_value or not activity:
            continue
        existing = db.query(models.WorkLog).filter(models.WorkLog.project_id == project_id,
                                                   models.WorkLog.date == date_value,
                                                   models.WorkLog.activity == activity).first()
        if existing:
            existing.area = row.get("area", existing.area)
            existing.pct = row.get("pct", existing.pct)
            existing.contractor = row.get("contractor", existing.contractor)
            existing.source_ref = source_ref
        else:
            db.add(models.WorkLog(project_id=project_id, date=date_value, activity=activity,
                                   area=row.get("area", ""), pct=row.get("pct", 0),
                                   contractor=row.get("contractor", ""), source_ref=source_ref))
        stored += 1
    return stored


def _upsert_milestones(db: Session, project_id: str, source_ref: str, rows: list[dict[str, Any]]) -> int:
    stored = 0
    for row in rows or []:
        title = row.get("title")
        date_value = row.get("date")
        if not title or not date_value:
            continue
        existing = db.query(models.Milestone).filter(models.Milestone.project_id == project_id,
                                                     models.Milestone.date == date_value,
                                                     models.Milestone.title == title).first()
        if existing:
            existing.status = row.get("status", existing.status)
            existing.detail = row.get("detail", existing.detail)
            existing.on_track = int(row.get("on_track", existing.on_track or 1))
            existing.source_ref = source_ref
        else:
            db.add(models.Milestone(project_id=project_id, date=date_value, title=title,
                                    status=row.get("status", "planned"), detail=row.get("detail", ""),
                                    on_track=int(row.get("on_track", 1)), source_ref=source_ref))
        stored += 1
    return stored


@router.post("/daily", response_model=DailyIngestResult)
async def ingest_daily(request: Request,
                       db: Session = Depends(get_db),
                       file: UploadFile | None = File(None),
                       project_id: str | None = Form(None),
                       source_ref: str = Form("")):
    payload = None
    if file is not None:
        parsed = _parse_upload_file(file)
        payload = DailyIngestPayload(project_id=project_id or "", source_ref=source_ref,
                                     bag_days=parsed["bag_days"], risks=parsed["risks"],
                                     report_snapshots=parsed["report_snapshots"],
                                     work_log=parsed["work_log"], milestones=parsed["milestones"])
    else:
        content_type = request.headers.get("content-type", "")
        if content_type.startswith("application/json"):
            payload = DailyIngestPayload(**await request.json())
        elif content_type.startswith("multipart/form-data"):
            payload = DailyIngestPayload(project_id=project_id or "", source_ref=source_ref)
        else:
            raise HTTPException(400, "Provide a JSON body or upload a file")
    if not portfolio.get_ops(payload.project_id):
        raise HTTPException(404, "Unknown project")
    counts = {
        "bag_days": _upsert_bag_days(db, payload.project_id, payload.source_ref, payload.bag_days),
        "risks": _upsert_risks(db, payload.project_id, payload.source_ref, payload.risks),
        "report_snapshots": _upsert_report_snapshots(db, payload.project_id, payload.source_ref, payload.report_snapshots),
        "work_log": _upsert_work_log(db, payload.project_id, payload.source_ref, payload.work_log),
        "milestones": _upsert_milestones(db, payload.project_id, payload.source_ref, payload.milestones),
    }
    db.commit()
    return DailyIngestResult(project_id=payload.project_id, source_ref=payload.source_ref, stored=counts)


@router.get("/status")
def status(db: Session = Depends(get_db)):
    return ingest_watch.status(db)


@router.post("/scan")
def scan():
    """Scan the watched drop-zone now and ingest new/changed files."""
    return ingest_watch.scan()


@router.get("/trend")
def trend(project_id: str, db: Session = Depends(get_db)):
    """Timestamped KPI snapshots for a project (change over time)."""
    rows = (db.query(models.Report)
            .filter(models.Report.project_id == project_id)
            .order_by(models.Report.generated_at).all())
    return [{"t": r.generated_at.isoformat(), **(r.data or {})} for r in rows]
